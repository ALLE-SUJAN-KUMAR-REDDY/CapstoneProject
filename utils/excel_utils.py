import openpyxl
from utils.logger import get_logger

logger = get_logger(__name__)


def read_test_data(file_path: str, sheet_name: str) -> list[dict]:
    """
    Reads all rows from the given Excel sheet and returns a list of dicts,
    where each dict maps column header -> cell value.
    """
    logger.info(f"Reading test data from '{file_path}', sheet='{sheet_name}'")
    wb = openpyxl.load_workbook(file_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError(f"Sheet '{sheet_name}' is empty.")

    headers = [str(h).strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        if any(cell is not None for cell in row):
            record = {headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))}
            data.append(record)

    logger.info(f"Loaded {len(data)} test record(s) from sheet '{sheet_name}'")
    return data
