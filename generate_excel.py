"""
generate_excel.py
-----------------
Utility script to regenerate the test_data.xlsx file.
Run this script whenever you need to reset or update test data:

    python generate_excel.py
"""

import openpyxl
import os


def generate_test_data(output_path: str = "test_data/test_data.xlsx") -> None:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    #  Sheet 1: One-Way Flight
    # ------------------------------------------------------------------ #
    ws1 = wb.active
    ws1.title = "OneWayFlight"

    headers = [
        "departure_city", "arrival_city", "departure_date",
        "flight_type", "flight_class", "passengers",
        "title", "first_name", "last_name",
        "email", "country_code", "phone",
        "nationality", "dob_day", "dob_month", "dob_year",
        "passport_number",
        "card_number", "card_expiry", "card_cvv", "card_name",
    ]
    ws1.append(headers)
    ws1.append([
        "BOM", "DEL", "17-03-2026",
        "One Way", "Economy", "1",
        "Mr", "John", "Doe",
        "ks7964@email.com", "IN +91", "9867543897",
        "India", "09", "09 Sep", "2004",
        "A8646348",
        "4242 4242 4242 4242", "12/28", "123", "John Doe",
    ])

    # ------------------------------------------------------------------ #
    #  Sheet 2: Round-Trip Flight
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet("RoundTripFlight")
    headers_rt = [
        "departure_city", "arrival_city", "departure_date", "return_date",
        "flight_type", "flight_class", "passengers",
        "title", "first_name", "last_name",
        "email", "country_code", "phone",
        "nationality", "dob_day", "dob_month", "dob_year",
        "passport_number",
        "card_number", "card_expiry", "card_cvv", "card_name",
    ]
    ws2.append(headers_rt)
    ws2.append([
        "BOM", "DEL", "17-03-2026", "24-03-2026",
        "Round Trip", "Economy", "1",
        "Mr", "Jane", "Smith",
        "jane.smith@email.com", "IN +91", "9876543210",
        "India", "15", "06 Jun", "1990",
        "B1234567",
        "4242 4242 4242 4242", "12/28", "123", "Jane Smith",
    ])

    # Style headers bold
    from openpyxl.styles import Font
    bold = Font(bold=True)
    for ws in [ws1, ws2]:
        for cell in ws[1]:
            cell.font = bold
        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(output_path)
    print(f"✅ Test data saved to: {output_path}")


if __name__ == "__main__":
    generate_test_data()
